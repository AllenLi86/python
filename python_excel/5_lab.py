from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('example.xlsx')
ws = wb.active

for row in range(2,16):
    for col in range(2, 12):
        char = get_column_letter(col)
        print(f'{ws[char+str(row)].value:2d}', end=" ")
    print("")
print("======")
for row in range(2, 17):
    char = 'L'
    ws[char+str(row)].value = f'=SUM(D{row}:F{row})'
    # print(ws[char+str(row)].value)
    
wb.save('example.xlsx')
print("saved!")

