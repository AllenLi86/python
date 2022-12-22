from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("new_excel.xlsx")
ws = wb.active

# row: 橫排
# col: 直排


# # 讀寫、修改資料
# for row in range(1, 5):
#     for col in range(1,6):
#         char = get_column_letter(col)
#         ws[char + str(row)].value = char + str(row)
        
# # 合併儲存格
# ws.merge_cells('A1:E1')
# ws.unmerge_cells('A1:E1')

# # 插入及刪除 行或列
# ws.insert_rows(3)
# ws.insert_cols(2)
# ws.delete_rows(4)
# ws.delete_cols(3)

# 移動資料
# ws.move_range('要移動的資料範圍', rows = 上(-)下(+)多少, cols = 左(-)右(+)多少)
ws.move_range('A3:E4', rows = 3, cols = 2)

        
wb.save('new_excel.xlsx')