from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


wb = load_workbook('example.xlsx', data_only=True)
ws = wb.active

for row in range(2, 17):
    char = 'L'
    print(ws[char+str(row)].value)