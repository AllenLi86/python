from openpyxl import Workbook, load_workbook

# 讀取檔案
wb = load_workbook('excel.xlsx')
# 要讀取的工作表
ws = wb['工作表2']

# 新增工作表
wb.create_sheet('test')

# 資料表名稱
print(wb.sheetnames)


wb.save('excel.xlsx')