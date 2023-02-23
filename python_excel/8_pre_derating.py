from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import re

wb = load_workbook('example.xlsx')
ws = wb["derating"]

j=1
while True:
    if ws[f'G{j}'].value is None:
        break
    else:
        j+=1
   
print("j= ", j)

for i in range(4, j):
    max = str(ws[f'G{i}'].value)
    maxNum = [float(s) for s in re.findall(r'-?\d+\.?\d*', max)]
    print(maxNum)

    app = str(ws[f'H{i}'].value)
    appNum = [float(s) for s in re.findall(r'-?\d+\.?\d*', app)]
    print(appNum)

    stress = f'VIN={round(appNum[0]/maxNum[0]*100)}%, IOUT={round(appNum[1]/maxNum[1]*100)}%, Tj={round(appNum[2]/maxNum[2]*100)}%'
    print(stress)
    ws[f'I{i}'] = stress

    state = f'{"Y" if round(appNum[0]/maxNum[0]*100) < 90 else "N"}, {"Y" if round(appNum[1]/maxNum[1]*100) < 80 else "N"}, {"Y" if appNum[2] < 110 else "N"}'
    ws[f'K{i}'] = state

    print("======")

wb.save("example.xlsx")
print("--- Done! ---")